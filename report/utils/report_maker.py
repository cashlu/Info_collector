"""
读取数据库，获取所有的报告记录。

"""
import os
import shutil
import openpyxl
from openpyxl.styles import Border, Side, Alignment

from report.models import Report, Town, Village

report_src = "d:/test/report_demo.xlsx"
reports_dest_dir = "d:/test/dest"
ledger_src = "d:/test/ledger_demo.xlsx"
ledger_dest_dir = "d:/test/ledger"


def report_file_gen(id, name, town, village):
    filename = '{}-{}-{}-{}.xlsx'.format(id, town, village, name)
    try:
        shutil.copyfile(report_src, os.path.join(reports_dest_dir, filename))
    except IOError as e:
        print(e)
        print('文件创建失败')
    else:
        print('{} 创建成功'.format(filename))
    return filename


def file_writer(dest_file, data_item):
    """
    将数据写入报告。
    """
    dest_book = openpyxl.load_workbook(dest_file)
    dest_sheet = dest_book['Sheet1']

    dest_sheet['H2'] = data_item["id"]
    dest_sheet['B4'] = data_item["name"]
    dest_sheet['H4'] = data_item["decade"]
    dest_sheet['B6'] = data_item["city"]
    dest_sheet['D6'] = data_item["town"]
    dest_sheet['E6'] = data_item["village"]
    # dest_sheet['B7'] = purpose
    # dest_sheet['B9'] = comment
    dest_sheet['A11'] = data_item["structure"]
    dest_sheet['B13'] = data_item["security"]
    dest_sheet['C13'] = data_item["yard2"]
    dest_sheet['E13'] = data_item["yard3"]
    dest_sheet['G13'] = data_item["yard4"]
    dest_sheet['B14'] = data_item["groundsill"]
    dest_sheet['C14'] = data_item["foundation2"]
    dest_sheet['E14'] = data_item["foundation3"]
    dest_sheet['G14'] = data_item["foundation4"]
    dest_sheet['B15'] = data_item["tilt"]
    dest_sheet['C15'] = data_item["house2"]
    dest_sheet['E15'] = data_item["house3"]
    dest_sheet['G15'] = data_item["house4"]
    dest_sheet['B16'] = data_item["upon"]
    dest_sheet['C16'] = data_item["top2"]
    dest_sheet['E16'] = data_item["top3"]
    dest_sheet['G16'] = data_item["top4"]
    dest_sheet['B17'] = data_item["fence"]
    dest_sheet['C17'] = data_item["wall2"]
    dest_sheet['E17'] = data_item["wall3"]
    dest_sheet['G17'] = data_item["wall4"]
    dest_sheet['B19'] = data_item["assess_level"]
    dest_sheet['C20'] = data_item["advice1"]
    dest_sheet['C21'] = data_item["advice2"]
    dest_sheet['B23'] = data_item["street_contract"]
    dest_sheet['E23'] = data_item["street_contract_phone"]
    # dest_sheet['B24'] = village_contact
    # dest_sheet['E24'] = village_mobile
    dest_sheet['B24'] = data_item["reviewer"]
    dest_sheet['E24'] = data_item["appraiser"]

    # 恢复样式
    dest_sheet['B8'] = '正房___间，'
    dest_sheet['D8'] = '共___层，'
    dest_sheet['E8'] = '附房___间'
    dest_sheet['B9'] = '倒塌___间，受损___间'

    replace_border(dest_sheet)
    dest_book.save(dest_file)


def replace_border(dest_sheet):
    """
    重新生成表格的边框样式。
    :param dest_sheet: 目标工作簿
    """
    left, right, top, bottom = [Side(style='thin', color='000000')] * 4
    dest_sheet['B3'].border = Border(top=top)
    dest_sheet['C3'].border = Border(top=top, bottom=bottom)
    dest_sheet['D3'].border = Border(top=top, bottom=bottom)
    dest_sheet['E3'].border = Border(top=top, bottom=bottom)
    dest_sheet['F3'].border = Border(top=top, bottom=bottom)
    dest_sheet['G3'].border = Border(top=top, bottom=bottom)
    dest_sheet['H3'].border = Border(top=top, bottom=bottom, right=right)
    dest_sheet['C4'].border = Border(top=top, bottom=bottom)
    dest_sheet['D4'].border = Border(top=top, bottom=bottom)
    dest_sheet['E4'].border = Border(top=top, bottom=bottom)
    dest_sheet['F4'].border = Border(top=top, bottom=bottom, right=right)
    dest_sheet['C5'].border = Border(top=top, bottom=bottom)
    dest_sheet['F5'].border = Border(top=top, bottom=bottom, right=right)
    dest_sheet['G5'].border = Border(bottom=bottom, left=left, right=right)
    dest_sheet['H5'].border = Border(bottom=bottom, right=right)
    dest_sheet['G6'].border = Border(top=top, bottom=bottom)
    dest_sheet['H6'].border = Border(top=top, bottom=bottom, right=right)
    dest_sheet['C7'].border = Border(top=top, bottom=bottom)
    dest_sheet['G7'].border = Border(top=top)
    dest_sheet['H7'].border = Border(top=top, right=right)
    dest_sheet['C9'].border = Border(top=top, bottom=bottom)
    dest_sheet['D9'].border = Border(top=top, bottom=bottom)
    dest_sheet['E9'].border = Border(top=top, bottom=bottom)
    dest_sheet['F9'].border = Border(top=top, bottom=bottom)
    dest_sheet['G9'].border = Border(top=top, bottom=bottom)
    dest_sheet['H9'].border = Border(top=top, bottom=bottom, right=right)
    dest_sheet['H10'].border = Border(right=right)
    dest_sheet['B11'].border = Border(top=top, bottom=bottom)
    dest_sheet['C11'].border = Border(top=top, bottom=bottom)
    dest_sheet['D11'].border = Border(top=top, bottom=bottom)
    dest_sheet['E11'].border = Border(top=top, bottom=bottom)
    dest_sheet['F11'].border = Border(top=top, bottom=bottom)
    dest_sheet['G11'].border = Border(top=top, bottom=bottom)
    dest_sheet['H11'].border = Border(top=top, bottom=bottom, right=right)
    dest_sheet['H12'].border = Border(right=right)
    dest_sheet['C11'].border = Border(top=top, bottom=bottom)
    dest_sheet['D11'].border = Border(top=top, bottom=bottom)
    dest_sheet['E11'].border = Border(top=top, bottom=bottom)
    dest_sheet['F11'].border = Border(top=top, bottom=bottom)
    dest_sheet['G11'].border = Border(top=top, bottom=bottom)
    dest_sheet['H11'].border = Border(top=top, bottom=bottom, right=right)
    dest_sheet['B13'].border = Border(top=top, bottom=bottom)
    dest_sheet['C13'].border = Border(top=top, bottom=bottom)
    dest_sheet['D13'].border = Border(top=top, bottom=bottom)
    dest_sheet['E13'].border = Border(top=top, bottom=bottom)
    dest_sheet['F13'].border = Border(top=top, bottom=bottom)
    dest_sheet['G13'].border = Border(top=top, bottom=bottom)
    dest_sheet['H13'].border = Border(top=top, bottom=bottom, right=right)
    dest_sheet['D14'].border = Border(top=top, bottom=bottom)
    dest_sheet['F14'].border = Border(top=top, bottom=bottom)
    dest_sheet['H14'].border = Border(top=top, bottom=bottom, right=right)
    dest_sheet['H15'].border = Border(right=right)
    dest_sheet['D16'].border = Border(top=top, bottom=bottom)
    dest_sheet['F16'].border = Border(top=top, bottom=bottom)
    dest_sheet['H16'].border = Border(top=top, bottom=bottom, right=right)
    dest_sheet['H17'].border = Border(right=right)
    dest_sheet['B18'].border = Border(top=top, bottom=bottom)
    dest_sheet['C18'].border = Border(top=top, bottom=bottom)
    dest_sheet['D18'].border = Border(top=top, bottom=bottom)
    dest_sheet['E18'].border = Border(top=top, bottom=bottom)
    dest_sheet['F18'].border = Border(top=top, bottom=bottom)
    dest_sheet['G18'].border = Border(top=top, bottom=bottom)
    dest_sheet['H18'].border = Border(top=top, bottom=bottom, right=right)
    dest_sheet['H19'].border = Border(right=right)
    dest_sheet['C20'].border = Border(top=top)
    dest_sheet['D20'].border = Border(top=top)
    dest_sheet['E20'].border = Border(top=top)
    dest_sheet['G20'].border = Border(top=top)
    dest_sheet['H20'].border = Border(top=top, right=right)
    dest_sheet['A21'].border = Border(bottom=bottom, right=right)
    dest_sheet['D21'].border = Border(bottom=bottom)
    dest_sheet['E21'].border = Border(bottom=bottom)
    dest_sheet['F21'].border = Border(bottom=bottom)
    dest_sheet['G21'].border = Border(bottom=bottom)
    dest_sheet['H21'].border = Border(bottom=bottom, right=right)


def data_preparer(field_name="", value=""):
    if field_name and value:
        queryset = Report.objects.filter(field_name=value)
    else:
        # 获取所有报告数据
        # queryset = Report.objects.all()
        queryset = Report.objects.exclude(id=412)

    data_list = []

    for item in queryset:
        print(item)
        data_item = {}
        data_item["id"] = item.id
        data_item["name"] = item.name
        data_item["identity"] = item.identity
        data_item["decade"] = item.get_decade_display()
        data_item["city"] = item.city.name
        data_item["town"] = item.town.name
        data_item["village"] = item.village.name
        data_item["purpose"] = item.get_purpose_display()
        data_item["comment"] = item.comment
        data_item["structure"] = item.get_structure_display()
        data_item["security"] = item.get_security_display()
        data_item["groundsill"] = item.get_groundsill_display()
        data_item["tilt"] = item.get_tilt_display()
        data_item["upon"] = item.get_upon_display()
        data_item["fence"] = item.get_fence_display()

        if item.security_detail.count() == 0:
            data_item["yard2"] = None
            data_item["yard3"] = None
            data_item["yard4"] = None
        elif item.security_detail.count() == 1:
            data_item["yard2"] = item.security_detail.all()[0].get_detail_display()
            data_item["yard3"] = None
            data_item["yard4"] = None
        elif item.security_detail.count() == 2:
            data_item["yard2"] = item.security_detail.all()[0].get_detail_display()
            data_item["yard3"] = item.security_detail.all()[1].get_detail_display()
            data_item["yard4"] = None
        elif item.security_detail.count() == 3:
            data_item["yard2"] = item.security_detail.all()[0].get_detail_display()
            data_item["yard3"] = item.security_detail.all()[1].get_detail_display()
            data_item["yard4"] = item.security_detail.all()[2].get_detail_display()

        # groundsill_detail = item.groundsill_detail.all()
        if item.groundsill_detail.count() == 0:
            data_item["foundation2"] = None
            data_item["foundation3"] = None
            data_item["foundation4"] = None
        elif item.groundsill_detail.count() == 1:
            data_item["foundation2"] = item.groundsill_detail.all()[0].get_detail_display()
            data_item["foundation3"] = None
            data_item["foundation4"] = None
        elif item.groundsill_detail.count() == 2:
            data_item["foundation2"] = item.groundsill_detail.all()[0].get_detail_display()
            data_item["foundation3"] = item.groundsill_detail.all()[1].get_detail_display()
            data_item["foundation4"] = None
        elif item.groundsill_detail.count() == 3:
            data_item["foundation2"] = item.groundsill_detail.all()[0].get_detail_display()
            data_item["foundation3"] = item.groundsill_detail.all()[1].get_detail_display()
            data_item["foundation4"] = item.groundsill_detail.all()[2].get_detail_display()

        # tilt_detail = item.tilt_detail.all()
        if item.tilt_detail.count() == 0:
            data_item["house2"] = None
            data_item["house3"] = None
            data_item["house4"] = None
        elif item.tilt_detail.count() == 1:
            data_item["house2"] = item.tilt_detail.all()[0].get_detail_display()
            data_item["house3"] = None
            data_item["house4"] = None
        elif item.tilt_detail.count() == 2:
            data_item["house2"] = item.tilt_detail.all()[0].get_detail_display()
            data_item["house3"] = item.tilt_detail.all()[1].get_detail_display()
            data_item["house4"] = None
        elif item.tilt_detail.count() == 3:
            data_item["house2"] = item.tilt_detail.all()[0].get_detail_display()
            data_item["house3"] = item.tilt_detail.all()[1].get_detail_display()
            data_item["house4"] = item.tilt_detail.all()[2].get_detail_display()

        # upon_detail = item.upon_detail.all()
        if item.upon_detail.count() == 0:
            data_item["top2"] = None
            data_item["top3"] = None
            data_item["top4"] = None
        elif item.upon_detail.count() == 1:
            data_item["top2"] = item.upon_detail.all()[0].get_detail_display()
            data_item["top3"] = None
            data_item["top4"] = None
        elif item.upon_detail.count() == 2:
            data_item["top2"] = item.upon_detail.all()[0].get_detail_display()
            data_item["top3"] = item.upon_detail.all()[1].get_detail_display()
            data_item["top4"] = None
        elif item.upon_detail.count() == 3:
            data_item["top2"] = item.upon_detail.all()[0].get_detail_display()
            data_item["top3"] = item.upon_detail.all()[1].get_detail_display()
            data_item["top4"] = item.upon_detail.all()[2].get_detail_display()

        # fence_detail = item.fence_detail.all()
        if item.fence_detail.count() == 0:
            data_item["wall2"] = None
            data_item["wall3"] = None
            data_item["wall4"] = None
        elif item.fence_detail.count() == 1:
            data_item["wall2"] = item.fence_detail.all()[0].get_detail_display()
            data_item["wall3"] = None
            data_item["wall4"] = None
        elif item.fence_detail.count() == 2:
            data_item["wall2"] = item.fence_detail.all()[0].get_detail_display()
            data_item["wall3"] = item.fence_detail.all()[1].get_detail_display()
            data_item["wall4"] = None
        elif item.fence_detail.count() == 3:
            data_item["wall2"] = item.fence_detail.all()[0].get_detail_display()
            data_item["wall3"] = item.fence_detail.all()[1].get_detail_display()
            data_item["wall4"] = item.fence_detail[2].get_detail_display()

        data_item["assess_level"] = item.get_assess_level_display()

        # advices = item.advice.all()
        if item.advice.count() == 0:
            data_item["advice1"] = None
            data_item["advice2"] = None
        elif item.advice.count() == 1:
            data_item["advice1"] = item.advice.all()[0].get_advice_display()
            data_item["advice2"] = None
        else:
            data_item["advice1"] = item.advice.all()[0].get_advice_display()
            data_item["advice2"] = item.advice.all()[1].get_advice_display()

        data_item["street_contract"] = item.street_contract
        data_item["street_contract_phone"] = item.street_contract_phone

        if item.reviewer:
            data_item["reviewer"] = item.reviewer.name
        else:
            data_item["reviewer"] = None
        data_item["created"] = item.created
        data_item["appraiser"] = item.appraiser.last_name + item.appraiser.first_name

        data_list.append(data_item)
    return data_list


def report_maker(data_list):
    for item in data_list:
        file = report_file_gen(item["id"], item["name"], item["town"], item["village"])
        full_path = os.path.join(reports_dest_dir, file)
        file_writer(full_path, item)


def make_report():
    data_list = data_preparer()
    report_maker(data_list)


def ledger_file_gen(city="", town="", village=""):
    filename = '{}{}{}受损农房.xlsx'.format(city, village, town)
    try:
        shutil.copyfile(ledger_src, os.path.join(ledger_dest_dir, filename))
    except IOError:
        print('文件创建失败')
    else:
        print('{} 创建成功'.format(filename))
    return filename


def make_ledger(level):
    if level == "city":
        # TODO: 市一级暂时没有实现
        pass
    elif level == "town":
        # town_list = []
        # for i in Report.objects.values("town"):
        #     town = Town.objects.get(id=i["town"])
        #     town_list.append(town.name)
        # town_set = list(set(town_list))
        # print(town_set)
        for town in Town.objects.all():
            # FIXME: 这里的层级结构硬编码了。需要修改
            filename = ledger_file_gen(city="潍坊市经济开发区", town=town)
            row_num = 3
            for report in Report.objects.filter(town=town):
                ledger_writer(report, filename, row_num, city="潍坊市经济开发区", town=town)
                row_num += 1
                print(report)

    elif level == "village":
        for village in Village.objects.all():
            # FIXME: 这里的层级结构硬编码了。需要修改
            filename = ledger_file_gen(city="潍坊市经济开发区", village=village)
            row_num = 3
            for report in Report.objects.filter(village=village):
                ledger_writer(report, filename, row_num, city="潍坊市经济开发区", village=village)
                row_num += 1
                print(report)


def ledger_writer(report, filename, row_num, city="", town="", village=""):
    fullname = os.path.join(ledger_dest_dir, filename)
    dest_book = openpyxl.load_workbook(fullname)
    dest_sheet = dest_book['Sheet1']
    left, right, top, bottom = [Side(style='thin', color='000000')] * 4
    dest_sheet['A1'] = filename[:-5]
    dest_sheet['A{}'.format(row_num)] = report.id
    dest_sheet.cell(row_num, 1).alignment = Alignment(horizontal='center',
                                                      vertical='center')
    dest_sheet['A{}'.format(row_num)].border = Border(top=top, bottom=bottom,
                                                      left=left, right=right)
    dest_sheet['B{}'.format(row_num)] = report.name
    dest_sheet.cell(row_num, 2).alignment = Alignment(horizontal='center',
                                                      vertical='center')
    dest_sheet['B{}'.format(row_num)].border = Border(top=top, bottom=bottom,
                                                      left=left, right=right)
    dest_sheet['C{}'.format(row_num)] = report.get_structure_display()
    dest_sheet.cell(row_num, 3).alignment = Alignment(horizontal='center',
                                                      vertical='center')
    dest_sheet['C{}'.format(row_num)].border = Border(top=top, bottom=bottom,
                                                      left=left, right=right)
    security = "场地安全程度" + report.get_security_display() + " : "
    for i in report.security_detail.all():
        security = security + i.get_detail_display() + "  "
    groundsill = "地基基础" + report.get_groundsill_display() + " : "
    for i in report.groundsill_detail.all():
        groundsill = groundsill + i.get_detail_display() + "  "

    tilt = "房屋整体倾斜" + report.get_tilt_display() + " : "
    for i in report.tilt_detail.all():
        tilt = tilt + i.get_detail_display() + "  "

    upon = "上部承重结构" + report.get_upon_display() + " : "
    for i in report.upon_detail.all():
        upon = upon + i.get_detail_display() + "  "

    fence = "围护结构" + report.get_fence_display() + " : "
    for i in report.fence_detail.all():
        fence = fence + i.get_detail_display() + "  "

    status = security + "\r\n" + groundsill + "\r\n" + tilt + "\r\n" + upon + "\r\n" + fence

    dest_sheet['D{}'.format(row_num)] = status
    dest_sheet.cell(row_num, 4).alignment = Alignment(wrapText=True,
                                                      # horizontal='center',
                                                      vertical='center')
    dest_sheet['D{}'.format(row_num)].border = Border(top=top, bottom=bottom,
                                                      left=left, right=right)
    dest_sheet['E{}'.format(row_num)] = report.get_assess_level_display()
    dest_sheet.cell(row_num, 5).alignment = Alignment(horizontal='center',
                                                      vertical='center')
    dest_sheet['E{}'.format(row_num)].border = Border(top=top, bottom=bottom,
                                                      left=left, right=right)
    dest_sheet['F{}'.format(row_num)] = report.comment
    dest_sheet.cell(row_num, 6).alignment = Alignment(horizontal='center',
                                                      vertical='center')
    dest_sheet['F{}'.format(row_num)].border = Border(top=top, bottom=bottom,
                                                      left=left, right=right)
    # row_num += 1
    dest_book.save(fullname)
